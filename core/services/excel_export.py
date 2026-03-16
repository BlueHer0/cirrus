"""CFDI → Excel/CSV export service.

Provides:
- export_cfdis_excel: QuerySet → .xlsx bytes (openpyxl)
- export_cfdis_csv: QuerySet → .csv bytes
- export_cfdi_detail_excel: Single CFDI XML → detailed .xlsx with conceptos + impuestos

For batch accounting exports with tax consolidation.
"""

import csv
import io
import logging
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("core.excel_export")

# Column definitions for the summary export
SUMMARY_COLUMNS = [
    ("UUID", 38),
    ("Fecha", 20),
    ("Tipo", 10),
    ("Serie", 10),
    ("Folio", 12),
    ("RFC Emisor", 15),
    ("Nombre Emisor", 35),
    ("RFC Receptor", 15),
    ("Nombre Receptor", 35),
    ("Subtotal", 16),
    ("IVA Trasl.", 14),
    ("ISR Ret.", 14),
    ("IVA Ret.", 14),
    ("Total", 16),
    ("Moneda", 8),
    ("Forma Pago", 12),
    ("Método Pago", 12),
    ("Estado SAT", 12),
]


# ── Style constants ───────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
MONEY_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=11)


def export_cfdis_excel(queryset, title: str = "CFDIs") -> bytes:
    """Export a CFDI queryset to a professional .xlsx file.

    Args:
        queryset: CFDI QuerySet (already filtered)
        title: Sheet title / header

    Returns:
        bytes of the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit

    # Header row
    for col_idx, (col_name, col_width) in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    money_cols = {10, 11, 12, 13, 14}  # Subtotal, IVA, ISR, IVA Ret, Total
    row_idx = 2

    for cfdi in queryset.iterator():
        row_data = [
            str(cfdi.uuid),
            cfdi.fecha.strftime("%Y-%m-%d %H:%M") if cfdi.fecha else "",
            cfdi.tipo_comprobante or "",
            cfdi.serie or "",
            cfdi.folio or "",
            cfdi.rfc_emisor or "",
            cfdi.nombre_emisor or "",
            cfdi.rfc_receptor or "",
            cfdi.nombre_receptor or "",
            float(cfdi.subtotal or 0),
            float(cfdi.total_impuestos_trasladados or 0),
            float(cfdi.isr_retenido or 0),
            float(cfdi.iva_retenido or 0),
            float(cfdi.total or 0),
            cfdi.moneda or "MXN",
            cfdi.forma_pago or "",
            cfdi.metodo_pago or "",
            cfdi.estado_sat or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in money_cols:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")

        row_idx += 1

    # Totals row
    if row_idx > 2:
        total_row = row_idx
        ws.cell(row=total_row, column=9, value="TOTALES:").font = SUBTOTAL_FONT

        for col_idx in money_cols:
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{row_idx - 1})",
            )
            cell.number_format = MONEY_FORMAT
            cell.font = SUBTOTAL_FONT
            cell.fill = SUBTOTAL_FILL
            cell.border = THIN_BORDER

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}{row_idx - 1}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_cfdis_csv(queryset) -> bytes:
    """Export a CFDI queryset to CSV bytes (UTF-8 BOM for Excel compatibility)."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([col_name for col_name, _ in SUMMARY_COLUMNS])

    # Data
    for cfdi in queryset.iterator():
        writer.writerow([
            str(cfdi.uuid),
            cfdi.fecha.strftime("%Y-%m-%d %H:%M") if cfdi.fecha else "",
            cfdi.tipo_comprobante or "",
            cfdi.serie or "",
            cfdi.folio or "",
            cfdi.rfc_emisor or "",
            cfdi.nombre_emisor or "",
            cfdi.rfc_receptor or "",
            cfdi.nombre_receptor or "",
            str(cfdi.subtotal or 0),
            str(cfdi.total_impuestos_trasladados or 0),
            str(cfdi.isr_retenido or 0),
            str(cfdi.iva_retenido or 0),
            str(cfdi.total or 0),
            cfdi.moneda or "MXN",
            cfdi.forma_pago or "",
            cfdi.metodo_pago or "",
            cfdi.estado_sat or "",
        ])

    # UTF-8 BOM for Excel auto-detection
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


def export_cfdi_detail_excel(xml_bytes: bytes) -> bytes:
    """Export a single CFDI XML to a detailed Excel with multiple sheets.

    Sheet 1: Comprobante (header + emisor/receptor)
    Sheet 2: Conceptos (line items with taxes)
    Sheet 3: Impuestos (tax summary)

    Args:
        xml_bytes: Raw XML bytes of the CFDI

    Returns:
        bytes of the .xlsx file
    """
    from sat_scrapper_core.cfdi_pdf import parse_cfdi_xml

    parsed = parse_cfdi_xml(xml_bytes)
    wb = Workbook()

    # ── Sheet 1: Comprobante ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Comprobante"

    comp = parsed.comprobante
    emisor = parsed.emisor
    receptor = parsed.receptor
    timbre = parsed.timbre

    info_rows = [
        ("UUID", timbre.get("uuid", "")),
        ("Versión", comp.get("version", "")),
        ("Fecha", str(comp.get("fecha", ""))),
        ("Tipo", f"{comp.get('tipo_comprobante', '')} - {comp.get('tipo_comprobante_desc', '')}"),
        ("Serie / Folio", f"{comp.get('serie', '')} {comp.get('folio', '')}".strip()),
        ("", ""),
        ("EMISOR", ""),
        ("RFC Emisor", emisor.get("rfc", "")),
        ("Nombre Emisor", emisor.get("nombre", "")),
        ("Régimen Fiscal", emisor.get("regimen_fiscal", "")),
        ("", ""),
        ("RECEPTOR", ""),
        ("RFC Receptor", receptor.get("rfc", "")),
        ("Nombre Receptor", receptor.get("nombre", "")),
        ("Uso CFDI", f"{receptor.get('uso_cfdi', '')} - {receptor.get('uso_cfdi_desc', '')}"),
        ("", ""),
        ("MONTOS", ""),
        ("Subtotal", float(comp.get("sub_total", 0))),
        ("Descuento", float(comp.get("descuento", 0) or 0)),
        ("Total", float(comp.get("total", 0))),
        ("Moneda", comp.get("moneda", "MXN")),
        ("Forma de Pago", f"{comp.get('forma_pago', '')} - {comp.get('forma_pago_desc', '')}"),
        ("Método de Pago", f"{comp.get('metodo_pago', '')} - {comp.get('metodo_pago_desc', '')}"),
    ]

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 50

    for row_idx, (label, value) in enumerate(info_rows, 1):
        label_cell = ws1.cell(row=row_idx, column=1, value=label)
        value_cell = ws1.cell(row=row_idx, column=2, value=value)

        if label in ("EMISOR", "RECEPTOR", "MONTOS"):
            label_cell.font = Font(bold=True, size=12, color="1F4E79")
        elif label:
            label_cell.font = Font(bold=True)

        if isinstance(value, float) and label in ("Subtotal", "Descuento", "Total"):
            value_cell.number_format = MONEY_FORMAT

    # ── Sheet 2: Conceptos ──────────────────────────────────────────
    ws2 = wb.create_sheet("Conceptos")
    concept_headers = [
        ("#", 5), ("Clave", 12), ("Descripción", 45), ("Unidad", 10),
        ("Cantidad", 12), ("P. Unitario", 16), ("Importe", 16),
        ("Descuento", 14), ("IVA", 14), ("ISR Ret.", 14),
    ]

    for col_idx, (header, width) in enumerate(concept_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.freeze_panes = "A2"

    for row_idx, concepto in enumerate(parsed.conceptos, 2):
        iva_amount = sum(
            float(t.get("importe", 0) or 0)
            for t in concepto.get("traslados", [])
            if t.get("impuesto") == "002"
        )
        isr_ret = sum(
            float(r.get("importe", 0) or 0)
            for r in concepto.get("retenciones", [])
            if r.get("impuesto") == "001"
        )

        row_data = [
            concepto.get("index", row_idx - 1),
            concepto.get("clave_prod_serv", ""),
            concepto.get("descripcion", ""),
            concepto.get("unidad", ""),
            float(concepto.get("cantidad", 0)),
            float(concepto.get("valor_unitario", 0)),
            float(concepto.get("importe", 0)),
            float(concepto.get("descuento", 0) or 0),
            iva_amount,
            isr_ret,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in {5, 6, 7, 8, 9, 10}:
                cell.number_format = MONEY_FORMAT

    # ── Sheet 3: Impuestos ──────────────────────────────────────────
    ws3 = wb.create_sheet("Impuestos")
    tax_headers = [("Tipo", 15), ("Impuesto", 12), ("Tasa/Cuota", 14), ("Base", 16), ("Importe", 16)]

    for col_idx, (header, width) in enumerate(tax_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    row_idx = 2
    impuestos = parsed.impuestos

    impuesto_names = {"001": "ISR", "002": "IVA", "003": "IEPS"}

    for traslado in impuestos.get("traslados", []):
        ws3.cell(row=row_idx, column=1, value="Traslado")
        ws3.cell(row=row_idx, column=2, value=impuesto_names.get(traslado.get("impuesto", ""), traslado.get("impuesto", "")))
        ws3.cell(row=row_idx, column=3, value=str(traslado.get("tasa", "")))
        ws3.cell(row=row_idx, column=4, value=float(traslado.get("base", 0) or 0)).number_format = MONEY_FORMAT
        ws3.cell(row=row_idx, column=5, value=float(traslado.get("importe", 0) or 0)).number_format = MONEY_FORMAT
        row_idx += 1

    for retencion in impuestos.get("retenciones", []):
        ws3.cell(row=row_idx, column=1, value="Retención")
        ws3.cell(row=row_idx, column=2, value=impuesto_names.get(retencion.get("impuesto", ""), retencion.get("impuesto", "")))
        ws3.cell(row=row_idx, column=3, value=str(retencion.get("tasa", "")))
        ws3.cell(row=row_idx, column=4, value=float(retencion.get("base", 0) or 0)).number_format = MONEY_FORMAT
        ws3.cell(row=row_idx, column=5, value=float(retencion.get("importe", 0) or 0)).number_format = MONEY_FORMAT
        row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
